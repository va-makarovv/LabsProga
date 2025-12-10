# python -m src.lab05.demo_lab05

from src.lab05.json_csv import *
from src.lab05.csv_xlsx import *

import json
import csv
import os


BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
DATA_SAMPLES = os.path.join(BASE_DIR, "data", "samples")
DATA_OUT = os.path.join(BASE_DIR, "data", "out")


def demo_json_to_csv():
    json_path = os.path.join(DATA_SAMPLES, "people.json")
    csv_path = os.path.join(DATA_OUT, "people_from_json.csv")

    json_to_csv(json_path, csv_path)
    print(f"[JSON→CSV] OK: {json_path} → {csv_path}")

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    print(f"[JSON→CSV] строк в CSV: {len(rows)}")
    if rows:
        print("[JSON→CSV] первая строка:", rows[0])


def demo_csv_to_json():
    csv_path = os.path.join(DATA_SAMPLES, "people.csv")
    json_path = os.path.join(DATA_OUT, "people_from_csv.json")

    csv_to_json(csv_path, json_path)
    print(f"[CSV→JSON] OK: {csv_path} → {json_path}")

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, list):
        print(f"[CSV→JSON] записей в JSON: {len(data)}")
        if data:
            print("[CSV→JSON] первая запись:", data[0])
    else:
        print("[CSV→JSON] тип данных:", type(data))
        print("[CSV→JSON] содержимое:", data)


def demo_csv_to_xlsx():
    csv_path = os.path.join(DATA_SAMPLES, "cities.csv")
    xlsx_path = os.path.join(DATA_OUT, "cities.xlsx")

    csv_to_xlsx(csv_path, xlsx_path)
    print(f"[CSV→XLSX] OK: {csv_path} → {xlsx_path}")

    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))

        print(f"[CSV→XLSX] строк : {len(rows)}")
        if rows:
            print("[CSV→XLSX] заголовок:", rows[0])
            if len(rows) > 1:
                print("[CSV→XLSX] первая строка данных:", rows[1])
    except ImportError:
        print("[CSV→XLSX] openpyxl не установлен")


if __name__ == "__main__":
    print("Демо lab05 ")
    demo_json_to_csv()
    print()
    demo_csv_to_json()
    print()
    demo_csv_to_xlsx()
