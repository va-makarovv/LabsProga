



from __future__ import annotations

import csv
from pathlib import Path


def _check_suffix(path: Path, expected: str) -> None:
    if path.suffix.lower() != expected:
        raise ValueError("Неверный тип файла")


def _ensure_parent_dir(path: Path) -> None:
    parent = path.parent
    if parent and not parent.exists():
        parent.mkdir(parents=True, exist_ok=True)


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:

    src = Path(csv_path)
    dst = Path(xlsx_path)
    _check_suffix(src, ".csv")
    _check_suffix(dst, ".xlsx")

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise ImportError(
            "Для конвертации в XLSX требуется пакет 'openpyxl'. Установите его или замените реализацию на xlsxwriter."
        ) from exc

    with src.open(encoding="utf-8") as f:
        sample = f.read(2048)
        if not sample.strip():
            raise ValueError("Пустой CSV или отсутствует заголовок")
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.get_dialect("excel")
        try:
            has_header = csv.Sniffer().has_header(sample)
        except csv.Error:
            has_header = True

        f.seek(0)
        reader = csv.reader(f, dialect=dialect)
        try:
            header = next(reader)
        except StopIteration:
            raise ValueError("Пустой CSV или отсутствует заголовок")

        if not has_header or not header or all((h or "").strip() == "" for h in header):
            raise ValueError("Пустой CSV или отсутствует заголовок")

        rows = [header]
        col_widths = [max(8, len(str(h))) for h in header]
        for row in reader:
            if len(row) < len(header):
                row = row + [""] * (len(header) - len(row))
            elif len(row) > len(header):
                row = row[: len(header)]
            rows.append(row)
            for idx, cell in enumerate(row):
                col_widths[idx] = max(col_widths[idx], len(str(cell)))

    _ensure_parent_dir(dst)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for row in rows:
        ws.append(["" if c is None else str(c) for c in row])

    for i, w in enumerate(col_widths, start=1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = max(8, w)

    wb.save(dst)